"""
xlsx_export.py — Export a saved StatMind session as a formatted Excel dashboard.

Input: the session snapshot the frontend already stores (saved_sessions.js):
    { capData: {col: capability_result}, spcData: {col: spc_result},
      normData: normality_result|null, grrData: {key: grr_result},
      capaReports: {key: capa_report}, globalProcType, globalFileName }

Output: an in-memory .xlsx (never written to disk — consistent with the
privacy policy's no-storage contract) containing:
    Dashboard   — verdict tiles + key metrics with semantic coloring
    Normality   — per-column test table
    Capability  — metrics + native histogram chart per column
    SPC         — control chart (native LineChart with CL/UCL/LCL) + alarms
    Gauge RR    — variance components + native bar chart
    CAPA        — matched rules, evidence, corrective/preventive actions
                  (the engine's generated narrative — the "AI inference" layer)

Design rules (regulated-industry semantics): red = out-of-spec/alarm only,
amber = marginal only, green = pass/in-control only. Arial throughout.
Engine-computed statistics are exported as values (they are the deliverable,
computed by the validated Python engines); simple aggregates over exported
data series use live Excel formulas.
"""
from __future__ import annotations

import datetime
import io
import re

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette (semantic — see frontend design tokens) ───────────────────────────
GREEN  = "2E7D32"; AMBER = "B26A00"; RED = "C62828"
GREEN_BG = "E8F5E9"; AMBER_BG = "FFF3E0"; RED_BG = "FFEBEE"
HEADER_BG = "1F2937"; SUBHEAD_BG = "EEF1F5"; BORDER = "D0D5DD"

F_TITLE  = Font(name="Arial", size=16, bold=True, color="FFFFFF")
F_H2     = Font(name="Arial", size=12, bold=True)
F_LABEL  = Font(name="Arial", size=10, color="475467")
F_BODY   = Font(name="Arial", size=10)
F_METRIC = Font(name="Arial", size=14, bold=True)
THIN = Border(*[Side(style="thin", color=BORDER)] * 4)


def _verdict_style(kind: str):
    return {
        "pass":     (Font(name="Arial", size=10, bold=True, color=GREEN), PatternFill("solid", start_color=GREEN_BG)),
        "marginal": (Font(name="Arial", size=10, bold=True, color=AMBER), PatternFill("solid", start_color=AMBER_BG)),
        "fail":     (Font(name="Arial", size=10, bold=True, color=RED),   PatternFill("solid", start_color=RED_BG)),
        "neutral":  (Font(name="Arial", size=10, bold=True),              PatternFill("solid", start_color=SUBHEAD_BG)),
    }[kind]


def _cpk_kind(cpk):
    if cpk is None: return "neutral"
    return "pass" if cpk >= 1.33 else "marginal" if cpk >= 1.0 else "fail"


def _grr_kind(pct):
    if pct is None: return "neutral"
    return "pass" if pct < 10 else "marginal" if pct <= 30 else "fail"


def _sev_kind(sev):
    return {"Critical": "fail", "Major": "marginal", "Minor": "pass"}.get(sev, "neutral")


def _safe_sheet_name(name: str, used: set) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", str(name))[:24] or "Sheet"
    candidate, i = base, 2
    while candidate in used:
        candidate = f"{base[:21]}_{i}"; i += 1
    used.add(candidate)
    return candidate


def _title_bar(ws, text, ncols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F_TITLE
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def _kv_row(ws, row, label, value, kind=None):
    lc = ws.cell(row=row, column=1, value=label); lc.font = F_LABEL
    vc = ws.cell(row=row, column=2, value=value); vc.font = F_BODY
    if kind:
        vc.font, vc.fill = _verdict_style(kind)
        vc.border = THIN
    return row + 1


def _table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = PatternFill("solid", start_color=SUBHEAD_BG)
        c.border = THIN
    return row + 1


def _num(v, nd=4):
    try:
        return round(float(v), nd)
    except (TypeError, ValueError):
        return v


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_dashboard(wb, session):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    _title_bar(ws, "StatMind Analysis Dashboard")
    for col, w in zip("ABCDEFGH", (26, 22, 16, 16, 16, 16, 16, 16)):
        ws.column_dimensions[col].width = w

    r = 3
    r = _kv_row(ws, r, "Source file", session.get("globalFileName") or "—")
    r = _kv_row(ws, r, "Process type", session.get("globalProcType") or "—")
    r = _kv_row(ws, r, "Exported", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    r = _kv_row(ws, r, "Generated by", "StatMind — statmind.tech")
    r += 1

    ws.cell(row=r, column=1, value="Verdicts").font = F_H2
    r += 1
    r = _table_header(ws, r, ["Analysis", "Parameter", "Key metric", "Value", "Verdict"])

    for col, cap in (session.get("capData") or {}).items():
        cpk = cap.get("cpk")
        kind = _cpk_kind(cpk)
        vals = ["Capability", col, "Cpk", _num(cpk, 3),
                cap.get("verdict") or {"pass": "Capable", "marginal": "Marginal", "fail": "Incapable"}.get(kind, "—")]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
        ws.cell(row=r, column=5).font, ws.cell(row=r, column=5).fill = _verdict_style(kind)
        r += 1

    for col, spc in (session.get("spcData") or {}).items():
        alarms = spc.get("total_alarms")
        in_ctl = spc.get("in_control")
        kind = "pass" if in_ctl else "fail"
        vals = ["SPC " + str(spc.get("chart_type") or ""), col, "Alarms", alarms,
                "In control" if in_ctl else f"Out of control ({alarms} alarms)"]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
        ws.cell(row=r, column=5).font, ws.cell(row=r, column=5).fill = _verdict_style(kind)
        r += 1

    norm = session.get("normData")
    norm_results = (norm or {}).get("results") if isinstance(norm, dict) else None
    if norm is not None and not norm_results and isinstance(norm, dict) and norm.get("column"):
        norm_results = [norm]  # single-column result stored directly
    for nres in (norm_results or []):
        verdict = nres.get("overall_verdict") or nres.get("verdict") or "—"
        kind = "pass" if "normal" == str(verdict).strip().lower() else \
               ("fail" if "non" in str(verdict).lower() else "neutral")
        vals = ["Normality", nres.get("column"), "Verdict", "", verdict]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
        ws.cell(row=r, column=5).font, ws.cell(row=r, column=5).fill = _verdict_style(kind)
        r += 1

    for key, grr in (session.get("grrData") or {}).items():
        pct = None
        g = grr.get("gauge_rr")
        if isinstance(g, dict):
            pct = g.get("pct_study_var")
        kind = _grr_kind(pct)
        vals = ["Gauge R&R", grr.get("column") or key, "%GRR (study var)", _num(pct, 2),
                grr.get("verdict") or "—"]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
        ws.cell(row=r, column=5).font, ws.cell(row=r, column=5).fill = _verdict_style(kind)
        r += 1

    for key, rep in (session.get("capaReports") or {}).items():
        sev = (rep.get("problem_statement") or {}).get("severity") or rep.get("severity")
        vals = ["CAPA", rep.get("fault_pattern") or rep.get("rule_id") or key,
                "Severity", "", sev or "—"]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
        ws.cell(row=r, column=5).font, ws.cell(row=r, column=5).fill = _verdict_style(_sev_kind(sev))
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="Semantic colors: green = pass/in-control · amber = marginal · red = out-of-spec/alarm."
            ).font = F_LABEL


def _build_normality(wb, session, used):
    norm = session.get("normData")
    if not norm:
        return
    ws = wb.create_sheet(_safe_sheet_name("Normality", used))
    ws.sheet_view.showGridLines = False
    _title_bar(ws, "Normality Tests", 7)
    for col, w in zip("ABCDEFG", (22, 20, 14, 14, 14, 12, 46)):
        ws.column_dimensions[col].width = w
    results = norm.get("results") if isinstance(norm, dict) else None
    if not results and isinstance(norm, dict) and norm.get("column"):
        results = [norm]
    r = 3
    for nres in (results or []):
        ws.cell(row=r, column=1, value=str(nres.get("column") or "")).font = F_H2
        r += 1
        r = _table_header(ws, r, ["Test", "Statistic", "p-value", "Critical", "α", "Reject H₀", "Interpretation"])
        for t in (nres.get("tests") or []):
            row_vals = [t.get("test_name"), _num(t.get("statistic"), 5), _num(t.get("p_value"), 5),
                        _num(t.get("critical_value"), 5), t.get("alpha"),
                        "Yes" if t.get("reject_null") else "No", t.get("interpretation")]
            for i, v in enumerate(row_vals):
                c = ws.cell(row=r, column=1 + i, value=v); c.font = F_BODY; c.border = THIN
            kind = "fail" if t.get("reject_null") else "pass"
            ws.cell(row=r, column=6).font, ws.cell(row=r, column=6).fill = _verdict_style(kind)
            r += 1
        verdict = nres.get("overall_verdict") or nres.get("verdict")
        if verdict:
            r = _kv_row(ws, r + 1, "Overall verdict", verdict,
                        "pass" if str(verdict).strip().lower() == "normal" else "fail")
        r += 1


def _build_capability(wb, session, used):
    for col, cap in (session.get("capData") or {}).items():
        ws = wb.create_sheet(_safe_sheet_name(f"Cap {col}", used))
        ws.sheet_view.showGridLines = False
        _title_bar(ws, f"Process Capability — {col}", 8)
        for c_, w in zip("ABCDEFGH", (20, 14, 20, 14, 14, 14, 14, 14)):
            ws.column_dimensions[c_].width = w
        r = 3
        pairs = [("n", cap.get("n")), ("Mean", _num(cap.get("mean"))),
                 ("Std (within)", _num(cap.get("std_within"))), ("Std (overall)", _num(cap.get("std_overall"))),
                 ("USL", cap.get("usl")), ("LSL", cap.get("lsl")), ("Target", cap.get("target")),
                 ("Cp", _num(cap.get("cp"), 3)), ("Cpk", _num(cap.get("cpk"), 3)),
                 ("Pp", _num(cap.get("pp"), 3)), ("Ppk", _num(cap.get("ppk"), 3)),
                 ("PPM (overall)", _num(cap.get("ppm_overall"), 1)), ("Sigma level", _num(cap.get("sigma_level"), 2))]
        for label, val in pairs:
            r = _kv_row(ws, r, label, val,
                        _cpk_kind(val) if label == "Cpk" else None)
        r = _kv_row(ws, r, "Verdict", cap.get("verdict") or "—", _cpk_kind(cap.get("cpk")))
        detail = cap.get("verdict_detail")
        if detail:
            r = _kv_row(ws, r, "Detail", detail)

        hist = cap.get("histogram_data") or {}
        # Engine schema: bin_centers + counts (bin_edges/bins kept as fallback)
        centers = hist.get("bin_centers") or []
        edges = hist.get("bin_edges") or hist.get("bins") or []
        counts = hist.get("counts") or hist.get("frequencies") or []
        if counts and not centers and len(edges) > len(counts):
            centers = [(float(edges[i]) + float(edges[i + 1])) / 2 for i in range(len(counts))]
        if counts and centers and len(centers) >= len(counts):
            hr = 3
            ws.cell(row=hr - 1, column=4, value="Histogram data").font = F_H2
            _table_header(ws, hr, ["Bin midpoint", "Count"], start_col=4)
            for i, cnt in enumerate(counts):
                ws.cell(row=hr + 1 + i, column=4, value=round(float(centers[i]), 3)).border = THIN
                ws.cell(row=hr + 1 + i, column=5, value=int(cnt)).border = THIN
            total_row = hr + 1 + len(counts)
            ws.cell(row=total_row, column=4, value="Total").font = F_LABEL
            ws.cell(row=total_row, column=5, value=f"=SUM(E{hr + 1}:E{hr + len(counts)})").font = F_BODY

            chart = BarChart()
            chart.type = "col"; chart.style = 10
            chart.title = f"{col} distribution"
            chart.y_axis.title = "Count"; chart.x_axis.title = col
            data = Reference(ws, min_col=5, min_row=hr, max_row=hr + len(counts))
            cats = Reference(ws, min_col=4, min_row=hr + 1, max_row=hr + len(counts))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            chart.width, chart.height = 18, 10
            ws.add_chart(chart, "G3")


def _build_spc(wb, session, used):
    for col, spc in (session.get("spcData") or {}).items():
        ws = wb.create_sheet(_safe_sheet_name(f"SPC {col}", used))
        ws.sheet_view.showGridLines = False
        _title_bar(ws, f"Control Chart ({spc.get('chart_type') or 'SPC'}) — {col}", 8)
        for c_, w in zip("ABCDEFGH", (12, 14, 12, 12, 12, 24, 14, 40)):
            ws.column_dimensions[c_].width = w

        vals = spc.get("primary_values") or []
        cl, ucl, lcl = spc.get("primary_cl"), spc.get("primary_ucl"), spc.get("primary_lcl")
        r = 3
        r = _table_header(ws, r, ["Subgroup", spc.get("primary_label") or "Value", "CL", "UCL", "LCL"])
        first_data = r
        for i, v in enumerate(vals):
            row_vals = [i + 1, _num(v), _num(cl), _num(ucl), _num(lcl)]
            for j, x in enumerate(row_vals):
                c = ws.cell(row=r, column=1 + j, value=x); c.font = F_BODY; c.border = THIN
            if ucl is not None and lcl is not None and v is not None and (v > ucl or v < lcl):
                ws.cell(row=r, column=2).font, ws.cell(row=r, column=2).fill = _verdict_style("fail")
            r += 1
        last_data = r - 1
        if vals:
            ws.cell(row=r, column=1, value="Mean").font = F_LABEL
            ws.cell(row=r, column=2, value=f"=AVERAGE(B{first_data}:B{last_data})").font = F_BODY

        if vals:
            chart = LineChart()
            chart.title = f"{col} — {spc.get('chart_type') or 'control chart'}"
            chart.style = 12
            chart.y_axis.title = spc.get("primary_label") or col
            chart.x_axis.title = "Subgroup"
            for c_idx, name in ((2, "Value"), (3, "CL"), (4, "UCL"), (5, "LCL")):
                ref = Reference(ws, min_col=c_idx, min_row=first_data - 1, max_row=last_data)
                s = Series(ref, title_from_data=True)
                if name != "Value":
                    s.graphicalProperties.line.dashStyle = "dash"
                    s.graphicalProperties.line.width = 12000
                s.smooth = False
                chart.series.append(s)
            cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
            chart.set_categories(cats)
            chart.width, chart.height = 24, 11
            ws.add_chart(chart, "G3")

        ar = max(r + 2, 3)
        ws.cell(row=ar, column=1, value="Alarms").font = F_H2
        ar += 1
        ar = _table_header(ws, ar, ["Rule", "Description", "Index"], start_col=1)
        alarms = (spc.get("western_electric_alarms") or []) + (spc.get("nelson_alarms") or [])
        for a in alarms[:60]:
            if not isinstance(a, dict):
                continue
            row_vals = [a.get("rule"), a.get("description"), a.get("index")]
            for j, x in enumerate(row_vals):
                c = ws.cell(row=ar, column=1 + j, value=x); c.font = F_BODY; c.border = THIN
            ws.cell(row=ar, column=1).font, ws.cell(row=ar, column=1).fill = _verdict_style("fail")
            ar += 1
        if not alarms:
            ws.cell(row=ar, column=1, value="No alarms — process in control")
            ws.cell(row=ar, column=1).font, ws.cell(row=ar, column=1).fill = _verdict_style("pass")


def _build_grr(wb, session, used):
    for key, grr in (session.get("grrData") or {}).items():
        ws = wb.create_sheet(_safe_sheet_name(f"GRR {grr.get('column') or key}", used))
        ws.sheet_view.showGridLines = False
        _title_bar(ws, f"Gauge R&R — {grr.get('column') or key}", 7)
        for c_, w in zip("ABCDEFG", (26, 14, 16, 16, 14, 14, 14)):
            ws.column_dimensions[c_].width = w
        r = 3
        r = _kv_row(ws, r, "Study", f"{grr.get('n_parts')} parts × {grr.get('n_operators')} operators × {grr.get('n_replicates')} trials")
        r = _kv_row(ws, r, "Method", grr.get("method"))
        if grr.get("tolerance") is not None:
            r = _kv_row(ws, r, "Tolerance", grr.get("tolerance"))
        r += 1
        r = _table_header(ws, r, ["Source", "Variance", "Std dev", "% Contribution", "% Study var"])
        first = r
        comps = [("Total Gauge R&R", grr.get("gauge_rr")),
                 ("  Repeatability (EV)", grr.get("repeatability")),
                 ("  Reproducibility (AV)", grr.get("reproducibility")),
                 ("  Operator × Part", grr.get("operator_by_part")),
                 ("Part-to-Part (PV)", grr.get("part_to_part"))]
        for name, comp in comps:
            if not isinstance(comp, dict):
                continue
            row_vals = [name, _num(comp.get("variance"), 5), _num(comp.get("std_dev"), 5),
                        _num(comp.get("pct_contribution"), 2), _num(comp.get("pct_study_var"), 2)]
            for j, x in enumerate(row_vals):
                c = ws.cell(row=r, column=1 + j, value=x); c.font = F_BODY; c.border = THIN
            r += 1
        last = r - 1
        pct = (grr.get("gauge_rr") or {}).get("pct_study_var") if isinstance(grr.get("gauge_rr"), dict) else None
        r = _kv_row(ws, r + 1, "Verdict", grr.get("verdict") or "—", _grr_kind(pct))
        if grr.get("verdict_detail"):
            r = _kv_row(ws, r, "Detail", grr.get("verdict_detail"))
        if grr.get("ndc") is not None:
            r = _kv_row(ws, r, "ndc (distinct categories)", grr.get("ndc"),
                        "pass" if (grr.get("ndc") or 0) >= 5 else "fail")

        if last >= first:
            chart = BarChart()
            chart.type = "col"; chart.style = 10
            chart.title = "% Study variation by source"
            data = Reference(ws, min_col=5, min_row=first - 1, max_row=last)
            cats = Reference(ws, min_col=1, min_row=first, max_row=last)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            chart.width, chart.height = 16, 9
            ws.add_chart(chart, "G3")


def _build_capa(wb, session, used):
    reports = session.get("capaReports") or {}
    if not reports:
        return
    ws = wb.create_sheet(_safe_sheet_name("CAPA", used))
    ws.sheet_view.showGridLines = False
    _title_bar(ws, "CAPA — Corrective & Preventive Actions", 6)
    for c_, w in zip("ABCDEF", (24, 60, 14, 16, 18, 20)):
        ws.column_dimensions[c_].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    r = 3
    for key, rep in reports.items():
        ps = rep.get("problem_statement") or {}
        sev = ps.get("severity") or rep.get("severity")
        ws.cell(row=r, column=1, value=rep.get("fault_pattern") or rep.get("rule_id") or key).font = F_H2
        sc = ws.cell(row=r, column=3, value=sev or "—")
        sc.font, sc.fill = _verdict_style(_sev_kind(sev))
        r += 1
        r = _kv_row(ws, r, "Rule", f"{rep.get('rule_id') or '—'} ({rep.get('process') or 'General'})")
        if rep.get("confidence_level"):
            r = _kv_row(ws, r, "Confidence", rep.get("confidence_level"))
        if ps.get("description"):
            c = ws.cell(row=r, column=1, value="Problem"); c.font = F_LABEL
            d = ws.cell(row=r, column=2, value=ps.get("description")); d.font = F_BODY; d.alignment = wrap
            r += 1
        for ev in (ps.get("statistical_evidence") or [])[:8]:
            c = ws.cell(row=r, column=1, value="Evidence"); c.font = F_LABEL
            d = ws.cell(row=r, column=2, value=str(ev)); d.font = F_BODY; d.alignment = wrap
            r += 1
        rca = rep.get("root_cause_analysis") or {}
        if rca.get("primary_hypothesis"):
            c = ws.cell(row=r, column=1, value="Primary root cause"); c.font = F_LABEL
            d = ws.cell(row=r, column=2, value=str(rca.get("primary_hypothesis"))); d.font = F_BODY; d.alignment = wrap
            r += 1
        actions = rep.get("corrective_actions") or []
        if actions:
            r = _table_header(ws, r + 1, ["Corrective action", "Detail", "Priority", "Timeline", "Owner", "Expected impact"])
            for a in actions:
                row_vals = ["", a.get("action"), a.get("priority"), a.get("timeline"),
                            a.get("owner"), a.get("expected_impact")]
                for j, x in enumerate(row_vals):
                    c = ws.cell(row=r, column=1 + j, value=x); c.font = F_BODY; c.border = THIN
                    c.alignment = wrap
                r += 1
        prevs = rep.get("preventive_actions") or []
        if prevs:
            r = _table_header(ws, r + 1, ["Preventive action", "Detail", "System change", "Timeline", "Owner", ""])
            for a in prevs:
                row_vals = ["", a.get("action"), a.get("system_change"), a.get("timeline"), a.get("owner"), ""]
                for j, x in enumerate(row_vals):
                    c = ws.cell(row=r, column=1 + j, value=x); c.font = F_BODY; c.border = THIN
                    c.alignment = wrap
                r += 1
        disp = rep.get("disposition") or {}
        if disp.get("recommendation"):
            kind = "fail" if disp["recommendation"] in ("Hold", "Scrap") else \
                   "pass" if disp["recommendation"] == "Release" else "marginal"
            r = _kv_row(ws, r + 1, "Disposition", disp.get("recommendation"), kind)
            if disp.get("rationale"):
                c = ws.cell(row=r, column=1, value="Rationale"); c.font = F_LABEL
                d = ws.cell(row=r, column=2, value=str(disp.get("rationale"))); d.font = F_BODY; d.alignment = wrap
                r += 1
        r += 2


# ── Public API ─────────────────────────────────────────────────────────────────

def build_session_workbook(session: dict) -> io.BytesIO:
    """Build the dashboard workbook fully in memory. Raises ValueError if the
    session contains no exportable analysis results."""
    if not isinstance(session, dict):
        raise ValueError("Session payload must be a JSON object.")
    has_any = any([session.get("capData"), session.get("spcData"),
                   session.get("normData"), session.get("grrData"),
                   session.get("capaReports")])
    if not has_any:
        raise ValueError("Session contains no analysis results to export. "
                         "Run at least one analysis, then export.")

    wb = Workbook()
    used: set = {"Dashboard"}
    _build_dashboard(wb, session)
    _build_normality(wb, session, used)
    _build_capability(wb, session, used)
    _build_spc(wb, session, used)
    _build_grr(wb, session, used)
    _build_capa(wb, session, used)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
