"""Tests for feat/session-excel-export: /api/v1/export/xlsx builds a valid
multi-sheet dashboard workbook from a session snapshot, fully in memory."""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
from xlsx_export import build_session_workbook

client = TestClient(main.app)


def _session():
    cap = {"column": "X", "n": 100, "cp": 1.2, "cpk": 1.05, "pp": 1.15, "ppk": 1.0,
           "mean": 450.0, "std_within": 8.0, "std_overall": 8.5, "usl": 480.0,
           "lsl": 420.0, "target": 450.0, "ppm_overall": 900.0, "sigma_level": 3.1,
           "verdict": "Marginal",
           "histogram_data": {"bin_centers": [440, 450, 460], "counts": [10, 60, 30]}}
    spc = {"chart_type": "Xbar-R", "column": "X", "in_control": False, "total_alarms": 2,
           "primary_values": [449.5, 450.2, 451.0, 458.9, 462.0],
           "primary_cl": 450.0, "primary_ucl": 456.0, "primary_lcl": 444.0,
           "primary_label": "Subgroup mean",
           "western_electric_alarms": [{"rule": "WE1", "description": "Beyond 3σ", "index": 4}],
           "nelson_alarms": []}
    grr = {"column": "X", "method": "ANOVA", "n_parts": 10, "n_operators": 3,
           "n_replicates": 3, "verdict": "Marginal", "ndc": 6,
           "gauge_rr": {"variance": 1.2, "std_dev": 1.1, "pct_contribution": 4.4, "pct_study_var": 21.0},
           "repeatability": {"variance": 1.0, "std_dev": 1.0, "pct_contribution": 3.7, "pct_study_var": 19.2},
           "reproducibility": {"variance": 0.2, "std_dev": 0.45, "pct_contribution": 0.7, "pct_study_var": 8.6},
           "operator_by_part": {"variance": 0.0, "std_dev": 0.0, "pct_contribution": 0.0, "pct_study_var": 0.0},
           "part_to_part": {"variance": 26.0, "std_dev": 5.1, "pct_contribution": 95.6, "pct_study_var": 97.8}}
    norm = {"results": [{"column": "X", "overall_verdict": "Normal",
                         "tests": [{"test_name": "Shapiro-Wilk", "statistic": 0.99,
                                    "p_value": 0.42, "critical_value": None,
                                    "alpha": 0.05, "reject_null": False,
                                    "interpretation": "No evidence against normality."}]}]}
    capa = {"rule_id": "GEN-003", "process": "General", "fault_pattern": "SPC Step Change",
            "confidence_level": "High",
            "problem_statement": {"severity": "Major", "description": "Step change detected",
                                  "statistical_evidence": ["2 WE alarms"]},
            "root_cause_analysis": {"primary_hypothesis": "Assignable cause event"},
            "corrective_actions": [{"action": "Investigate", "priority": "P1",
                                    "timeline": "24h", "owner": "PE", "expected_impact": "Contain"}],
            "preventive_actions": [{"action": "Add OCAP", "system_change": "SPC plan",
                                    "timeline": "2w", "owner": "QE"}],
            "disposition": {"recommendation": "Hold", "rationale": "OOC lot"}}
    return {"capData": {"X": cap}, "spcData": {"X": spc}, "grrData": {"X": grr},
            "normData": norm, "capaReports": {"k": capa},
            "globalProcType": "Etch", "globalFileName": "demo.csv"}


def test_export_endpoint_returns_valid_workbook():
    r = client.post("/api/v1/export/xlsx", json={"session": _session(), "name": "My Study"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert 'filename="My Study.xlsx"' in r.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[0] == "Dashboard"
    assert any(s.startswith("Cap") for s in wb.sheetnames)
    assert any(s.startswith("SPC") for s in wb.sheetnames)
    assert any(s.startswith("GRR") for s in wb.sheetnames)
    assert "CAPA" in wb.sheetnames and "Normality" in wb.sheetnames


def test_export_contains_native_charts():
    buf = build_session_workbook(_session())
    wb = load_workbook(buf)
    cap_sheet = next(s for s in wb.sheetnames if s.startswith("Cap"))
    spc_sheet = next(s for s in wb.sheetnames if s.startswith("SPC"))
    grr_sheet = next(s for s in wb.sheetnames if s.startswith("GRR"))
    assert len(wb[cap_sheet]._charts) == 1   # histogram
    assert len(wb[spc_sheet]._charts) == 1   # control chart
    assert len(wb[grr_sheet]._charts) == 1   # variance components


def test_export_dashboard_verdicts_present():
    wb = load_workbook(build_session_workbook(_session()))
    text = " ".join(str(c.value) for row in wb["Dashboard"].iter_rows() for c in row if c.value)
    for expected in ("Capability", "SPC", "Gauge R&R", "CAPA", "Marginal", "demo.csv"):
        assert expected in text


def test_export_empty_session_400():
    r = client.post("/api/v1/export/xlsx", json={"session": {}, "name": "x"})
    assert r.status_code == 400
    assert "no analysis results" in r.json()["detail"]


def test_export_partial_session_ok():
    """Only SPC present — no crash on missing sections."""
    s = {"spcData": _session()["spcData"]}
    r = client.post("/api/v1/export/xlsx", json={"session": s, "name": "spc only"})
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert not any(x.startswith("Cap ") for x in wb.sheetnames)


def test_export_filename_sanitized():
    r = client.post("/api/v1/export/xlsx",
                    json={"session": {"spcData": _session()["spcData"]},
                          "name": '../..\\evil":<name>'})
    assert r.status_code == 200
    fname = r.headers["content-disposition"].split('filename="')[1]
    # Security property: no path separators, quotes, or leading dots survive
    for bad in ("/", "\\", '"', "<", ">", ":"):
        assert bad not in fname.rstrip('"')
    assert not fname.startswith(".")


def test_export_weird_column_names_safe_sheet_titles():
    s = {"spcData": {"Rate [nm/min]: *final?": _session()["spcData"]["X"]}}
    r = client.post("/api/v1/export/xlsx", json={"session": s, "name": "w"})
    assert r.status_code == 200
    load_workbook(io.BytesIO(r.content))   # openpyxl would reject invalid titles
